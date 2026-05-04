import requests
import json
import time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
import os


class SamsungCrawler:
    def __init__(self, excel_file: str = "samsung_product_bulk.xlsx"):
        self.base_url = "https://searchapi.samsung.com/v6/front/b2c/product/finder/global"
        self.excel_file = excel_file
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'en-US,en;q=0.9,vi;q=0.8',
            'Accept-Encoding': 'gzip, deflate, br',
            'Referer': 'https://www.samsung.com/',
            'Origin': 'https://www.samsung.com',
            'Connection': 'keep-alive',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-site',
        })
        
    def set_cookies_from_string(self, cookie_string: str):
        """
        Set cookies từ string (copy từ browser)
        Format: "name1=value1; name2=value2; ..."
        """
        cookies = {}
        for item in cookie_string.split(';'):
            item = item.strip()
            if '=' in item:
                key, value = item.split('=', 1)
                cookies[key.strip()] = value.strip()
        
        # Set cookies vào session
        for name, value in cookies.items():
            self.session.cookies.set(name, value, domain='.samsung.com')
        print(f"Đã set {len(cookies)} cookies")
    
    def fetch_page(self, page: int = 1, page_size: int = 60) -> Dict[str, Any]:
        """
        Fetch một trang dữ liệu từ API Samsung
        """
        # Thử với các params khác nhau
        params = {
            'onlyFilterInfoYN': 'N'
        }
        
        # Thêm pagination nếu cần
        if page > 1:
            params['page'] = page
            params['pageSize'] = page_size
        else:
            # Thử với pageSize ngay từ đầu
            params['pageSize'] = page_size
            params['page'] = page
        
        try:
            response = self.session.get(self.base_url, params=params, timeout=30)
            print(f"Status code: {response.status_code}")
            if response.status_code != 200:
                print(f"Response text: {response.text[:500]}")
                # Nếu vẫn 403, thử không có pageSize
                if response.status_code == 403 and 'pageSize' in params:
                    print("Thử lại không có pageSize...")
                    params.pop('pageSize', None)
                    params.pop('page', None)
                    response = self.session.get(self.base_url, params=params, timeout=30)
                    print(f"Status code (retry): {response.status_code}")
            
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            print(f"Lỗi khi fetch trang {page}: {e}")
            if hasattr(e, 'response') and e.response is not None:
                print(f"Response status: {e.response.status_code}")
            return None
    
    def extract_product_data(self, product: Dict[str, Any]) -> Dict[str, Any]:
        """
        Extract và map dữ liệu từ product JSON sang format Excel
        """
        # Lấy modelList, nếu không có thì dùng dict rỗng
        model_list = product.get('modelList', [])
        first_model = model_list[0] if model_list else {}
        
        # Lấy keySummary
        key_summary = product.get('keySummary', [])
        
        # Xử lý galleryImageAlt (tags)
        gallery_image_alt = product.get('galleryImageAlt', '')
        
        # Xử lý priceDisplay
        price_display = product.get('priceDisplay', '')
        
        # Xử lý originPdpUrl (slug)
        origin_pdp_url = product.get('originPdpUrl', '')
        # Loại bỏ https://www.samsung.com nếu có
        if origin_pdp_url.startswith('https://www.samsung.com'):
            slug = origin_pdp_url.replace('https://www.samsung.com', '').strip('/')
        else:
            slug = origin_pdp_url.strip('/')
        
        # Xử lý thumbUrl - lấy tất cả từ tất cả models
        thumb_urls = []
        for model in model_list:
            thumb_url = model.get('thumbUrl', '')
            if thumb_url:
                # Thêm https: nếu chưa có
                if thumb_url.startswith('//'):
                    thumb_url = 'https:' + thumb_url
                elif not thumb_url.startswith('http'):
                    thumb_url = 'https://' + thumb_url
                thumb_urls.append(thumb_url)
        thumbnail_img = ','.join(thumb_urls) if thumb_urls else ''
        
        # Xử lý galleryImage - lấy tất cả từ tất cả models
        gallery_images = []
        for model in model_list:
            gallery_image = model.get('galleryImage', [])
            if isinstance(gallery_image, list):
                for img in gallery_image:
                    if img:
                        # Thêm https: nếu chưa có
                        if img.startswith('//'):
                            img = 'https:' + img
                        elif not img.startswith('http'):
                            img = 'https://' + img
                        gallery_images.append(img)
            elif gallery_image:
                if gallery_image.startswith('//'):
                    gallery_image = 'https:' + gallery_image
                elif not gallery_image.startswith('http'):
                    gallery_image = 'https://' + gallery_image
                gallery_images.append(gallery_image)
        photos = ','.join(gallery_images) if gallery_images else ''
        
        # Xử lý keySummary description
        descriptions = []
        for summary in key_summary:
            desc = summary.get('description', '')
            if desc:
                descriptions.append(desc)
        meta_description = ','.join(descriptions) if descriptions else ''
        
        # Lấy modelCode từ model đầu tiên
        sku = first_model.get('modelCode', '') if first_model else ''
        
        # Tạo dữ liệu product
        product_data = {
            'name': product.get('fmyMarketingName', ''),
            'description': '',
            'category_id': 55,
            'multi_categories': '9, 55',
            'brand_id': 24,
            'video_provider': 'youtube',
            'video_link': 'youtube',
            'tags': gallery_image_alt,
            'unit_price': price_display,
            'unit': 'Chiếc',
            'slug': slug,
            'current_stock': 100,
            'est_shipping_days': 5,
            'sku': sku,
            'meta_title': product.get('displayName', ''),
            'meta_description': meta_description,
            'thumbnail_img': thumbnail_img,
            'photos': photos
        }
        
        return product_data
    
    def crawl_all_products(self) -> List[Dict[str, Any]]:
        """
        Crawl tất cả sản phẩm từ API với pagination
        """
        all_products = []
        page = 1
        page_size = 60  # Có thể điều chỉnh
        
        print("Bắt đầu crawl dữ liệu từ API Samsung...")
        
        while True:
            print(f"Đang fetch trang {page}...")
            data = self.fetch_page(page, page_size)
            
            if not data:
                print(f"Không thể lấy dữ liệu trang {page}. Dừng lại.")
                break
            
            # Kiểm tra cấu trúc response
            # Có thể là data.response.resultData.productList hoặc data.productList
            product_list = None
            if 'response' in data and 'resultData' in data['response']:
                product_list = data['response']['resultData'].get('productList', [])
            elif 'resultData' in data:
                product_list = data['resultData'].get('productList', [])
            elif 'productList' in data:
                product_list = data['productList']
            
            if not product_list or len(product_list) == 0:
                print(f"Không còn sản phẩm nào ở trang {page}. Kết thúc.")
                break
            
            print(f"Tìm thấy {len(product_list)} sản phẩm ở trang {page}")
            
            # Extract dữ liệu từ mỗi product
            for product in product_list:
                try:
                    product_data = self.extract_product_data(product)
                    all_products.append(product_data)
                except Exception as e:
                    print(f"Lỗi khi extract dữ liệu product: {e}")
                    continue
            
            # Kiểm tra xem còn trang nào không
            # Có thể kiểm tra totalCount hoặc số lượng products
            total_count = None
            if 'response' in data and 'resultData' in data['response']:
                total_count = data['response']['resultData'].get('totalCount', 0)
            elif 'resultData' in data:
                total_count = data['resultData'].get('totalCount', 0)
            elif 'totalCount' in data:
                total_count = data.get('totalCount', 0)
            
            if total_count:
                print(f"Tổng số sản phẩm: {total_count}, đã lấy: {len(all_products)}")
                if len(all_products) >= total_count:
                    break
            
            # Nếu số lượng products ít hơn page_size, có thể đã hết
            if len(product_list) < page_size:
                print("Đã lấy hết tất cả sản phẩm.")
                break
            
            page += 1
            time.sleep(1)  # Delay để tránh rate limiting
        
        print(f"\nHoàn thành! Tổng cộng đã crawl {len(all_products)} sản phẩm.")
        return all_products
    
    def save_to_excel(self, products: List[Dict[str, Any]]):
        """
        Lưu dữ liệu vào file Excel
        """
        # Định nghĩa các cột
        columns = [
            'name', 'description', 'category_id', 'multi_categories', 'brand_id',
            'video_provider', 'video_link', 'tags', 'unit_price', 'unit',
            'slug', 'current_stock', 'est_shipping_days', 'sku', 'meta_title',
            'meta_description', 'thumbnail_img', 'photos'
        ]
        
        # Kiểm tra xem file Excel đã tồn tại chưa
        if os.path.exists(self.excel_file):
            wb = load_workbook(self.excel_file)
            ws = wb.active
        else:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            # Ghi header
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = col_name
                cell.font = cell.font.copy(bold=True)
        
        # Tìm dòng cuối cùng có dữ liệu
        last_row = ws.max_row
        if last_row == 1 and ws.cell(row=1, column=1).value is None:
            start_row = 1
            # Ghi header nếu chưa có
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = col_name
                cell.font = cell.font.copy(bold=True)
            start_row = 2
        else:
            start_row = last_row + 1
        
        # Ghi dữ liệu
        print(f"\nĐang ghi {len(products)} sản phẩm vào Excel...")
        for row_idx, product in enumerate(products, start=start_row):
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = product.get(col_name, '')
                # Xử lý các giá trị đặc biệt
                if isinstance(value, (list, dict)):
                    value = str(value)
                cell.value = value
        
        # Auto-adjust column widths
        for col_idx in range(1, len(columns) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if row.value:
                        max_length = max(max_length, len(str(row.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(self.excel_file)
        print(f"Đã lưu dữ liệu vào file: {self.excel_file}")


def main():
    crawler = SamsungCrawler()
    
    # Hướng dẫn người dùng set cookies nếu cần
    print("=" * 60)
    print("HƯỚNG DẪN:")
    print("1. Mở browser và vào https://www.samsung.com")
    print("2. Mở DevTools (F12) > Network tab")
    print("3. Tìm request đến searchapi.samsung.com")
    print("4. Copy cookies từ request headers")
    print("5. Paste vào đây (hoặc Enter để bỏ qua và thử không cookies)")
    print("=" * 60)
    
    cookie_input = input("\nNhập cookies (hoặc Enter để bỏ qua): ").strip()
    if cookie_input:
        crawler.set_cookies_from_string(cookie_input)
    
    # Crawl tất cả sản phẩm
    products = crawler.crawl_all_products()
    
    if products:
        # Lưu vào Excel
        crawler.save_to_excel(products)
        print(f"\n✅ Hoàn thành! Đã crawl và lưu {len(products)} sản phẩm vào Excel.")
    else:
        print("\n❌ Không có dữ liệu để lưu.")
        print("\n💡 Gợi ý:")
        print("   - Thử mở browser và copy cookies từ Network tab")
        print("   - Hoặc sử dụng script crawl_samsung_selenium.py với Selenium")


if __name__ == "__main__":
    main()

