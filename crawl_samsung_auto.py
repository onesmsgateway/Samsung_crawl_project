import json
import time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
import requests


class SamsungCrawlerAuto:
    def __init__(self, excel_file: str = "samsung_product_bulk.xlsx", target_url: str = "https://www.samsung.com/vn/smartphones/all-smartphones/"):
        self.excel_file = excel_file
        self.target_url = target_url
        self.api_url = "https://searchapi.samsung.com/v6/front/b2c/product/finder/global"
        self.all_products = []
        self.driver = None
        self.session = requests.Session()
        self.processed_request_ids = set()  # Track các request đã xử lý
        
    def setup_driver(self):
        """Setup Chrome driver với performance logging"""
        chrome_options = Options()
        chrome_options.add_argument('--headless=new') # Chạy headless để không cần mở cửa sổ UI
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        chrome_options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        
        # Enable performance logging để bắt network requests
        chrome_options.set_capability('goog:loggingPrefs', {'performance': 'ALL'})
        
        driver_path = ChromeDriverManager().install()
        # Fix lỗi WinError 193 khi webdriver-manager lấy nhầm file license/notice
        if not driver_path.endswith('.exe'):
            parent_dir = os.path.dirname(driver_path)
            potential_exe = os.path.join(parent_dir, 'chromedriver.exe')
            if os.path.exists(potential_exe):
                driver_path = potential_exe
            else:
                # Thử tìm trong thư mục con nếu có (cấu trúc mới của chromedriver)
                for root, dirs, files in os.walk(parent_dir):
                    if 'chromedriver.exe' in files:
                        driver_path = os.path.join(root, 'chromedriver.exe')
                        break
        
        print(f"Using ChromeDriver at: {driver_path}")
        service = Service(driver_path)
        self.driver = webdriver.Chrome(service=service, options=chrome_options)
        return self.driver
    
    def get_cookies_from_browser(self):
        """Lấy cookies từ browser và set vào requests session"""
        cookies = self.driver.get_cookies()
        for cookie in cookies:
            self.session.cookies.set(cookie['name'], cookie['value'], domain=cookie.get('domain', '.samsung.com'))
        print(f"Got {len(cookies)} cookies from browser")
    
    def fetch_api_with_javascript(self, page: int = 1, page_size: int = 60):
        """Gọi API trực tiếp từ browser bằng JavaScript"""
        url = f"{self.api_url}?onlyFilterInfoYN=N&page={page}&pageSize={page_size}"
        
        try:
            result = self.driver.execute_async_script("""
                var callback = arguments[arguments.length - 1];
                var url = arguments[0];
                fetch(url, {
                    method: 'GET',
                    headers: {
                        'Accept': 'application/json, text/plain, */*',
                        'Referer': 'https://www.samsung.com/',
                    },
                    credentials: 'include'
                })
                .then(response => response.json())
                .then(data => {
                    callback(JSON.stringify(data));
                })
                .catch(error => {
                    callback(JSON.stringify({error: error.toString()}));
                });
            """, url)
            
            if result:
                data = json.loads(result)
                if 'error' in data:
                    print(f"JavaScript error: {data['error']}")
                    return None
                return data
            return None
        except Exception as e:
            print(f"Error executing JavaScript: {e}")
            return None
    
    def extract_product_data(self, product: Dict[str, Any]) -> Dict[str, Any]:
        """Extract và map dữ liệu từ product JSON sang format Excel"""
        model_list = product.get('modelList', [])
        key_summary = product.get('keySummary', [])
        gallery_image_alt = product.get('galleryImageAlt', '')
        
        # Tìm model đầu tiên có giá trị (cho priceDisplay, originPdpUrl, keySummary)
        first_model_with_data = None
        for model in model_list:
            if model:  # Nếu model có dữ liệu
                first_model_with_data = model
                break
        
        # Nếu không tìm thấy, dùng model đầu tiên (có thể rỗng)
        first_model = first_model_with_data if first_model_with_data else (model_list[0] if model_list else {})
        
        # Lấy promotionPrice - ƯU TIÊN promotionPrice
        price_display = ''
        for model in model_list:
            if not model:
                continue
            # Ưu tiên promotionPrice (theo yêu cầu)
            price_display = model.get('promotionPrice', '')
            # Nếu không có promotionPrice, thử priceDisplay
            if not price_display:
                price_display = model.get('priceDisplay', '')
            # Nếu vẫn không có, thử các trường khác
            if not price_display:
                price_display = (model.get('afterTaxPriceDisplay', '') or
                               model.get('rrpPriceDisplay', '') or
                               model.get('lowestWasPriceDisplay', '') or
                               model.get('price', ''))
            if price_display:
                break  # Đã tìm thấy giá, dừng lại
        
        # Nếu vẫn không có, thử từ root level
        if not price_display:
            price_display = product.get('promotionPrice', '') or product.get('priceDisplay', '') or product.get('price', '')
        
        # Convert sang string nếu là số
        if price_display and not isinstance(price_display, str):
            price_display = str(price_display)
        
        # Loại bỏ "VND" và các ký tự không cần thiết khỏi giá
        if price_display:
            price_display = str(price_display).replace('VND', '').replace('vnd', '').replace('₫', '').replace('đ', '').strip()
            # Loại bỏ tất cả dấu chấm và phẩy (để chỉ còn số thuần túy)
            price_display = price_display.replace('.', '').replace(',', '').strip()
        
        # Nếu không có giá, điền giá trị mặc định là 1 (cho tablets)
        if not price_display:
            price_display = '1'
        
        # Lấy description từ modelList[0].keySummary - lấy TẤT CẢ description, cách nhau bởi dấu phẩy
        description = ''
        for model in model_list:
            if not model:
                continue
            model_key_summary = model.get('keySummary', [])
            if model_key_summary and len(model_key_summary) > 0:
                descriptions = []
                for summary in model_key_summary:
                    desc = summary.get('description', '')
                    if desc:
                        descriptions.append(desc)
                if descriptions:
                    description = ', '.join(descriptions)
                    break  # Đã tìm thấy description, dừng lại
        
        # Lấy slug từ modelList[0].originPdpUrl (nếu phần tử đầu tiên ko có thì lấy phần tử tiếp theo có giá trị)
        origin_pdp_url = ''
        for model in model_list:
            if not model:
                continue
            origin_pdp_url = model.get('originPdpUrl', '')
            if origin_pdp_url:
                break  # Đã tìm thấy, dừng lại
        
        # Nếu không có trong modelList, thử từ root level
        if not origin_pdp_url:
            origin_pdp_url = product.get('originPdpUrl', '')
        
        # Xử lý slug
        if origin_pdp_url.startswith('https://www.samsung.com'):
            slug = origin_pdp_url.replace('https://www.samsung.com', '').strip('/')
        else:
            slug = origin_pdp_url.strip('/')
        
        # Xử lý thumbUrl - chỉ lấy 1 giá trị gặp lần đầu thôi
        thumbnail_img = ''
        for model in model_list:
            if not model:
                continue
            thumb_url = model.get('thumbUrl', '')
            if thumb_url:
                if thumb_url.startswith('//'):
                    thumb_url = 'https:' + thumb_url
                elif not thumb_url.startswith('http'):
                    thumb_url = 'https://' + thumb_url
                thumbnail_img = thumb_url
                break  # Chỉ lấy giá trị đầu tiên
        
        # Xử lý galleryImage
        gallery_images = []
        for model in model_list:
            gallery_image = model.get('galleryImage', [])
            if isinstance(gallery_image, list):
                for img in gallery_image:
                    if img:
                        if img.startswith('//'):
                            img = 'https:' + img
                        elif not img.startswith('http'):
                            img = 'https://' + img
                        gallery_images.append(img)
            elif gallery_image:
                if gallery_image.startswith('//'):
                    gallery_image = 'https:' + gallery_image
                elif not gallery_image.startswith('http'):
                    gallery_image = 'https://' + gallery_image
                gallery_images.append(gallery_image)
        photos = ','.join(gallery_images) if gallery_images else ''
        
        # Xử lý keySummary description
        descriptions = []
        for summary in key_summary:
            desc = summary.get('description', '')
            if desc:
                descriptions.append(desc)
        meta_description = ','.join(descriptions) if descriptions else ''
        
        # Lấy SKU từ model đầu tiên có giá trị
        sku = ''
        for model in model_list:
            if not model:
                continue
            sku = model.get('modelCode', '')
            if sku:
                break  # Đã tìm thấy, dừng lại
        
        product_data = {
            'name': product.get('fmyMarketingName', ''),
            'description': description,
            'category_id': 55,
            'multi_categories': '9, 55',
            'brand_id': 24,
            'video_provider': 'youtube',
            'video_link': 'youtube',
            'tags': gallery_image_alt,
            'unit_price': price_display,
            'unit': 'Cái',
            'slug': slug,
            'current_stock': 100,
            'est_shipping_days': 5,
            'sku': sku,
            'meta_title': product.get('displayName', ''),
            'meta_description': meta_description,
            'thumbnail_img': thumbnail_img,
            'photos': photos
        }
        
        return product_data
    
    def extract_network_responses(self):
        """Extract API responses từ browser performance logs"""
        products_data = []
        try:
            logs = self.driver.get_log('performance')
            # Lưu response đầu tiên để debug (chỉ 1 lần)
            if not hasattr(self, '_debug_saved'):
                self._debug_saved = True
                for log in logs[:50]:  # Chỉ check 50 logs đầu
                    try:
                        log_message = json.loads(log['message'])
                        message = log_message.get('message', {})
                        if message.get('method') == 'Network.responseReceived':
                            response = message.get('params', {}).get('response', {})
                            url = response.get('url', '')
                            if 'searchapi.samsung.com' in url and 'onlyFilterInfoYN=N' in url:
                                request_id = message.get('params', {}).get('requestId', '')
                                try:
                                    response_body = self.driver.execute_cdp_cmd('Network.getResponseBody', {'requestId': request_id})
                                    if response_body and 'body' in response_body:
                                        body_text = response_body['body']
                                        if response_body.get('base64Encoded', False):
                                            import base64
                                            body_text = base64.b64decode(body_text).decode('utf-8')
                                        # Lưu vào file để debug
                                        with open('debug_response.json', 'w', encoding='utf-8') as f:
                                            f.write(body_text)
                                        print("Saved sample response to debug_response.json")
                                        break
                                except:
                                    pass
                    except:
                        continue
            
            for log in logs:
                try:
                    log_message = json.loads(log['message'])
                    message = log_message.get('message', {})
                    method = message.get('method', '')
                    
                    if method == 'Network.responseReceived':
                        response = message.get('params', {}).get('response', {})
                        url = response.get('url', '')
                        
                        if 'searchapi.samsung.com' in url and 'onlyFilterInfoYN=N' in url:
                            request_id = message.get('params', {}).get('requestId', '')
                            
                            # Bỏ qua nếu đã xử lý request này
                            if request_id in self.processed_request_ids:
                                continue
                            
                            try:
                                # Lấy response body
                                response_body = self.driver.execute_cdp_cmd('Network.getResponseBody', {'requestId': request_id})
                                if response_body and 'body' in response_body:
                                    body_text = response_body['body']
                                    if response_body.get('base64Encoded', False):
                                        import base64
                                        body_text = base64.b64decode(body_text).decode('utf-8')
                                    
                                    data = json.loads(body_text)
                                    self.processed_request_ids.add(request_id)  # Đánh dấu đã xử lý
                                    products_data.append(data)
                            except Exception as e:
                                continue
                except:
                    continue
        except Exception as e:
            print(f"Error extracting network logs: {e}")
        
        return products_data
    
    def crawl_all_products(self):
        """Crawl tất cả sản phẩm bằng cách scroll và bắt network requests"""
        print("Opening browser...")
        # Mở trang target URL
        print(f"Opening: {self.target_url}")
        self.driver.get(self.target_url)
        time.sleep(8)  # Đợi trang load và xử lý cookies
        
        # Đợi một chút để trang load xong
        try:
            WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
        except:
            pass
        
        print("Scrolling page to trigger API calls...")
        last_height = self.driver.execute_script("return document.body.scrollHeight")
        scroll_count = 0
        max_scrolls = 100
        no_new_products_count = 0
        
        def find_and_click_load_more():
            """Tìm và click nút 'Xem thêm' nếu có"""
            try:
                # Tìm các nút có thể là "Xem thêm"
                selectors = [
                    "//button[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'xem thêm')]",
                    "//button[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'load more')]",
                    "//a[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'xem thêm')]",
                    "//button[contains(@class, 'load-more')]",
                    "//button[contains(@class, 'show-more')]",
                    "//*[contains(@class, 'load-more')]//button",
                ]
                
                for selector in selectors:
                    try:
                        elements = self.driver.find_elements(By.XPATH, selector)
                        for element in elements:
                            if element.is_displayed() and element.is_enabled():
                                # Scroll đến element
                                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", element)
                                time.sleep(1)
                                element.click()
                                print("Clicked 'Load more' button")
                                time.sleep(3)  # Đợi sản phẩm load
                                return True
                    except:
                        continue
            except Exception as e:
                pass
            return False
        
        while scroll_count < max_scrolls:
            products_before = len(self.all_products)
            
            # Tìm và click nút "Xem thêm" trước khi scroll
            if scroll_count % 3 == 0:  # Mỗi 3 lần scroll thì thử click
                find_and_click_load_more()
            
            # Scroll từ từ, từng phần nhỏ để trigger API calls tự nhiên hơn
            scroll_step = 300  # Scroll mỗi lần 300px
            current_position = self.driver.execute_script("return window.pageYOffset;")
            max_scroll = self.driver.execute_script("return document.body.scrollHeight - window.innerHeight;")
            
            # Scroll từ từ trong 5 bước
            for step in range(5):
                scroll_to = min(current_position + scroll_step * (step + 1), max_scroll)
                self.driver.execute_script(f"window.scrollTo({{top: {scroll_to}, behavior: 'smooth'}});")
                time.sleep(0.8)  # Đợi scroll animation
                
                # Extract network responses sau mỗi lần scroll nhỏ
                responses = self.extract_network_responses()
                for response in responses:
                    product_list = None
                    if 'response' in response and 'resultData' in response['response']:
                        product_list = response['response']['resultData'].get('productList', [])
                    elif 'resultData' in response:
                        product_list = response['resultData'].get('productList', [])
                    elif 'productList' in response:
                        product_list = response['productList']
                    
                    if product_list:
                        for product in product_list:
                            try:
                                # Lọc bỏ các sản phẩm không phải item
                                product_name = product.get('fmyMarketingName', '')
                                display_name = product.get('displayName', '')
                                
                                # Loại bỏ "Pre-Registration eVoucher" và "Galaxy S25+ (chỉ có tại Samsung.com)"
                                if 'Pre-Registration eVoucher' in str(product_name) or 'Pre-Registration eVoucher' in str(display_name):
                                    continue
                                if 'Galaxy S25+' in str(product_name) and 'chỉ có tại Samsung.com' in str(product_name):
                                    continue
                                if 'Galaxy S25+' in str(display_name) and 'chỉ có tại Samsung.com' in str(display_name):
                                    continue
                                
                                product_data = self.extract_product_data(product)
                                
                                # Đảm bảo có giá (nếu không có sẽ là '1')
                                if not product_data.get('unit_price'):
                                    product_data['unit_price'] = '1'
                                
                                sku = product_data.get('sku', '')
                                if sku and not any(p.get('sku') == sku for p in self.all_products):
                                    self.all_products.append(product_data)
                                elif not sku:
                                    slug = product_data.get('slug', '')
                                    if slug and not any(p.get('slug') == slug for p in self.all_products):
                                        self.all_products.append(product_data)
                            except Exception as e:
                                continue
            
            # Scroll đến cuối để đảm bảo
            self.driver.execute_script("window.scrollTo({top: document.body.scrollHeight, behavior: 'smooth'});")
            time.sleep(3)  # Đợi API response
            
            # Extract network responses một lần nữa sau khi scroll xong
            responses = self.extract_network_responses()
            for response in responses:
                # Extract product list từ response
                product_list = None
                if 'response' in response and 'resultData' in response['response']:
                    product_list = response['response']['resultData'].get('productList', [])
                elif 'resultData' in response:
                    product_list = response['resultData'].get('productList', [])
                elif 'productList' in response:
                    product_list = response['productList']
                
                if product_list:
                    print(f"Found {len(product_list)} products in response")
                    for product in product_list:
                        try:
                            # Lọc bỏ các sản phẩm không phải item
                            product_name = product.get('fmyMarketingName', '')
                            display_name = product.get('displayName', '')
                            
                            # Loại bỏ "Pre-Registration eVoucher" và "Galaxy S25+ (chỉ có tại Samsung.com)"
                            if 'Pre-Registration eVoucher' in str(product_name) or 'Pre-Registration eVoucher' in str(display_name):
                                continue
                            if 'Galaxy S25+' in str(product_name) and 'chỉ có tại Samsung.com' in str(product_name):
                                continue
                            if 'Galaxy S25+' in str(display_name) and 'chỉ có tại Samsung.com' in str(display_name):
                                continue
                            
                            product_data = self.extract_product_data(product)
                            
                            # Đảm bảo có giá (nếu không có sẽ là '1')
                            if not product_data.get('unit_price'):
                                product_data['unit_price'] = '1'
                            
                            # Kiểm tra duplicate bằng SKU
                            sku = product_data.get('sku', '')
                            if sku and not any(p.get('sku') == sku for p in self.all_products):
                                self.all_products.append(product_data)
                            elif not sku:
                                # Nếu không có SKU, kiểm tra bằng slug
                                slug = product_data.get('slug', '')
                                if slug and not any(p.get('slug') == slug for p in self.all_products):
                                    self.all_products.append(product_data)
                        except Exception as e:
                            print(f"Error extracting product data: {e}")
                            continue
            
            # Kiểm tra xem có sản phẩm mới không
            products_after = len(self.all_products)
            if products_after == products_before:
                no_new_products_count += 1
                if no_new_products_count >= 3:
                    print("No new products after 3 scrolls. Stopping.")
                    break
            else:
                no_new_products_count = 0
            
            # Thử click "Xem thêm" một lần nữa sau khi scroll
            if find_and_click_load_more():
                time.sleep(3)
            
            # Kiểm tra xem có scroll thêm được không
            new_height = self.driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                print("Scrolled to end of page.")
                # Thử click "Xem thêm" một lần nữa
                if find_and_click_load_more():
                    time.sleep(4)
                    new_height = self.driver.execute_script("return document.body.scrollHeight")
                
                if new_height == last_height:
                    # Đợi thêm một chút xem có sản phẩm mới load không
                    time.sleep(3)
                    new_height = self.driver.execute_script("return document.body.scrollHeight")
                    if new_height == last_height:
                        break
            last_height = new_height
            scroll_count += 1
            print(f"Scrolled {scroll_count} times, found {len(self.all_products)} products...")
        
        print(f"\nDone! Total crawled {len(self.all_products)} products.")
    
    def save_to_excel(self, products: List[Dict[str, Any]]):
        """Lưu dữ liệu vào file Excel"""
        columns = [
            'name', 'description', 'category_id', 'multi_categories', 'brand_id',
            'video_provider', 'video_link', 'tags', 'unit_price', 'unit',
            'slug', 'current_stock', 'est_shipping_days', 'sku', 'meta_title',
            'meta_description', 'thumbnail_img', 'photos'
        ]
        
        if os.path.exists(self.excel_file):
            wb = load_workbook(self.excel_file)
            ws = wb.active
        else:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = col_name
                cell.font = cell.font.copy(bold=True)
        
        last_row = ws.max_row
        if last_row == 1 and ws.cell(row=1, column=1).value is None:
            start_row = 1
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = col_name
                cell.font = cell.font.copy(bold=True)
            start_row = 2
        else:
            start_row = last_row + 1
        
        print(f"\nWriting {len(products)} products to Excel...")
        for row_idx, product in enumerate(products, start=start_row):
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = product.get(col_name, '')
                if isinstance(value, (list, dict)):
                    value = str(value)
                cell.value = value
        
        for col_idx in range(1, len(columns) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if row.value:
                        max_length = max(max_length, len(str(row.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(self.excel_file)
        print(f"Data saved to file: {self.excel_file}")
    
    def crawl(self):
        """Main crawl method"""
        try:
            self.setup_driver()
            self.crawl_all_products()
            
            if self.all_products:
                self.save_to_excel(self.all_products)
                print(f"\nDone! Crawled and saved {len(self.all_products)} products to Excel.")
            else:
                print("\nNo data to save.")
        finally:
            if self.driver:
                self.driver.quit()


def main():
    import sys
    
    # Kiểm tra xem có tham số command line không
    if len(sys.argv) > 1:
        if sys.argv[1] == 'tablets':
            # Crawl tablets
            crawler = SamsungCrawlerAuto(
                excel_file="samsung_tab_product_bulk.xlsx",
                target_url="https://www.samsung.com/vn/tablets/all-tablets/"
            )
        elif sys.argv[1] == 'watches':
            # Crawl watches
            crawler = SamsungCrawlerAuto(
                excel_file="samsung_watch_product_bulk.xlsx",
                target_url="https://www.samsung.com/vn/watches/all-watches/"
            )
        elif sys.argv[1] == 'audio' or sys.argv[1] == 'buds':
            # Crawl audio-sound
            crawler = SamsungCrawlerAuto(
                excel_file="samsung_buds_product_bulk.xlsx",
                target_url="https://www.samsung.com/vn/audio-sound/all-audio-sound/"
            )
        elif sys.argv[1] == 'rings' or sys.argv[1] == 'ring':
            # Crawl rings
            crawler = SamsungCrawlerAuto(
                excel_file="samsung_ring_product_bulk.xlsx",
                target_url="https://www.samsung.com/vn/rings/all-rings/"
            )
        elif sys.argv[1] == 'accessories' or sys.argv[1] == 'mobile-accessories':
            # Crawl mobile-accessories
            crawler = SamsungCrawlerAuto(
                excel_file="samsung_accessories_product_bulk.xlsx",
                target_url="https://www.samsung.com/vn/mobile-accessories/all-mobile-accessories/"
            )
        else:
            # Crawl smartphones (mặc định)
            crawler = SamsungCrawlerAuto()
    else:
        # Crawl smartphones (mặc định)
        crawler = SamsungCrawlerAuto()
    
    crawler.crawl()


if __name__ == "__main__":
    main()

