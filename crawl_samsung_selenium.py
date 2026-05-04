import json
import time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager


class SamsungCrawlerSelenium:
    def __init__(self, excel_file: str = "samsung_product_bulk.xlsx"):
        self.excel_file = excel_file
        self.api_url_pattern = "searchapi.samsung.com/v6/front/b2c/product/finder/global"
        self.all_products = []
        self.driver = None
        
    def setup_driver(self):
        """Setup Chrome driver với DevTools Protocol để intercept network requests"""
        chrome_options = Options()
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        chrome_options.add_argument('user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        
        # Enable performance logging để bắt network requests
        chrome_options.set_capability('goog:loggingPrefs', {'performance': 'ALL'})
        
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=chrome_options)
        return self.driver
    
    def extract_network_responses(self) -> List[Dict[str, Any]]:
        """Extract các network responses từ browser logs"""
        products_data = []
        logs = self.driver.get_log('performance')
        
        for log in logs:
            try:
                log_message = json.loads(log['message'])
                message = log_message.get('message', {})
                method = message.get('method', '')
                
                # Bắt response từ API Samsung
                if method == 'Network.responseReceived':
                    response = message.get('params', {}).get('response', {})
                    url = response.get('url', '')
                    
                    if self.api_url_pattern in url and 'onlyFilterInfoYN=N' in url:
                        request_id = message.get('params', {}).get('requestId', '')
                        # Lấy response body
                        try:
                            response_body = self.driver.execute_cdp_cmd('Network.getResponseBody', {'requestId': request_id})
                            if response_body and 'body' in response_body:
                                body_text = response_body['body']
                                # Decode base64 nếu cần
                                if response_body.get('base64Encoded', False):
                                    import base64
                                    body_text = base64.b64decode(body_text).decode('utf-8')
                                
                                data = json.loads(body_text)
                                products_data.append(data)
                        except Exception as e:
                            print(f"Lỗi khi lấy response body: {e}")
            except Exception as e:
                continue
        
        return products_data
    
    def crawl_with_scroll(self, max_scrolls: int = 10):
        """Mở trang Samsung và scroll để trigger API calls"""
        print("Đang mở trang Samsung...")
        
        # Mở một trang có sản phẩm Samsung (ví dụ trang search hoặc products)
        self.driver.get("https://www.samsung.com/us/")
        time.sleep(3)
        
        # Tìm và click vào menu Products hoặc tìm kiếm
        try:
            # Thử tìm search box hoặc products menu
            search_selectors = [
                "input[type='search']",
                "input[name='search']",
                ".search-input",
                "#search",
                "[aria-label*='search' i]"
            ]
            
            search_found = False
            for selector in search_selectors:
                try:
                    search_element = WebDriverWait(self.driver, 5).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                    )
                    search_element.click()
                    time.sleep(1)
                    search_found = True
                    break
                except:
                    continue
            
            if not search_found:
                print("Không tìm thấy search box, thử scroll trang...")
        except Exception as e:
            print(f"Không thể tìm search box: {e}")
        
        print("Đang scroll trang để trigger API calls...")
        last_height = self.driver.execute_script("return document.body.scrollHeight")
        scroll_count = 0
        
        while scroll_count < max_scrolls:
            # Scroll xuống
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)  # Đợi API response
            
            # Extract network responses sau mỗi lần scroll
            responses = self.extract_network_responses()
            for response in responses:
                self.process_api_response(response)
            
            # Kiểm tra xem có scroll thêm được không
            new_height = self.driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                print("Đã scroll đến cuối trang.")
                break
            last_height = new_height
            scroll_count += 1
            print(f"Đã scroll {scroll_count} lần, tìm thấy {len(self.all_products)} sản phẩm...")
    
    def process_api_response(self, data: Dict[str, Any]):
        """Xử lý response từ API và extract products"""
        # Kiểm tra cấu trúc response
        product_list = None
        if 'response' in data and 'resultData' in data['response']:
            product_list = data['response']['resultData'].get('productList', [])
        elif 'resultData' in data:
            product_list = data['resultData'].get('productList', [])
        elif 'productList' in data:
            product_list = data['productList']
        
        if product_list:
            print(f"Tìm thấy {len(product_list)} sản phẩm trong response")
            for product in product_list:
                try:
                    product_data = self.extract_product_data(product)
                    # Kiểm tra duplicate
                    if not any(p.get('sku') == product_data.get('sku') for p in self.all_products if product_data.get('sku')):
                        self.all_products.append(product_data)
                except Exception as e:
                    print(f"Lỗi khi extract dữ liệu product: {e}")
                    continue
    
    def extract_product_data(self, product: Dict[str, Any]) -> Dict[str, Any]:
        """Extract và map dữ liệu từ product JSON sang format Excel"""
        # Lấy modelList, nếu không có thì dùng dict rỗng
        model_list = product.get('modelList', [])
        first_model = model_list[0] if model_list else {}
        
        # Lấy keySummary
        key_summary = product.get('keySummary', [])
        
        # Xử lý galleryImageAlt (tags)
        gallery_image_alt = product.get('galleryImageAlt', '')
        
        # Xử lý priceDisplay
        price_display = product.get('priceDisplay', '')
        
        # Xử lý originPdpUrl (slug)
        origin_pdp_url = product.get('originPdpUrl', '')
        # Loại bỏ https://www.samsung.com nếu có
        if origin_pdp_url.startswith('https://www.samsung.com'):
            slug = origin_pdp_url.replace('https://www.samsung.com', '').strip('/')
        else:
            slug = origin_pdp_url.strip('/')
        
        # Xử lý thumbUrl - lấy tất cả từ tất cả models
        thumb_urls = []
        for model in model_list:
            thumb_url = model.get('thumbUrl', '')
            if thumb_url:
                # Thêm https: nếu chưa có
                if thumb_url.startswith('//'):
                    thumb_url = 'https:' + thumb_url
                elif not thumb_url.startswith('http'):
                    thumb_url = 'https://' + thumb_url
                thumb_urls.append(thumb_url)
        thumbnail_img = ','.join(thumb_urls) if thumb_urls else ''
        
        # Xử lý galleryImage - lấy tất cả từ tất cả models
        gallery_images = []
        for model in model_list:
            gallery_image = model.get('galleryImage', [])
            if isinstance(gallery_image, list):
                for img in gallery_image:
                    if img:
                        # Thêm https: nếu chưa có
                        if img.startswith('//'):
                            img = 'https:' + img
                        elif not img.startswith('http'):
                            img = 'https://' + img
                        gallery_images.append(img)
            elif gallery_image:
                if gallery_image.startswith('//'):
                    gallery_image = 'https:' + gallery_image
                elif not gallery_image.startswith('http'):
                    gallery_image = 'https://' + gallery_image
                gallery_images.append(gallery_image)
        photos = ','.join(gallery_images) if gallery_images else ''
        
        # Xử lý keySummary description
        descriptions = []
        for summary in key_summary:
            desc = summary.get('description', '')
            if desc:
                descriptions.append(desc)
        meta_description = ','.join(descriptions) if descriptions else ''
        
        # Lấy modelCode từ model đầu tiên
        sku = first_model.get('modelCode', '') if first_model else ''
        
        # Tạo dữ liệu product
        product_data = {
            'name': product.get('fmyMarketingName', ''),
            'description': '',
            'category_id': 55,
            'multi_categories': '9, 55',
            'brand_id': 24,
            'video_provider': 'youtube',
            'video_link': 'youtube',
            'tags': gallery_image_alt,
            'unit_price': price_display,
            'unit': 'Chiếc',
            'slug': slug,
            'current_stock': 100,
            'est_shipping_days': 5,
            'sku': sku,
            'meta_title': product.get('displayName', ''),
            'meta_description': meta_description,
            'thumbnail_img': thumbnail_img,
            'photos': photos
        }
        
        return product_data
    
    def save_to_excel(self, products: List[Dict[str, Any]]):
        """Lưu dữ liệu vào file Excel"""
        # Định nghĩa các cột
        columns = [
            'name', 'description', 'category_id', 'multi_categories', 'brand_id',
            'video_provider', 'video_link', 'tags', 'unit_price', 'unit',
            'slug', 'current_stock', 'est_shipping_days', 'sku', 'meta_title',
            'meta_description', 'thumbnail_img', 'photos'
        ]
        
        # Kiểm tra xem file Excel đã tồn tại chưa
        if os.path.exists(self.excel_file):
            wb = load_workbook(self.excel_file)
            ws = wb.active
        else:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            # Ghi header
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = col_name
                cell.font = cell.font.copy(bold=True)
        
        # Tìm dòng cuối cùng có dữ liệu
        last_row = ws.max_row
        if last_row == 1 and ws.cell(row=1, column=1).value is None:
            start_row = 1
            # Ghi header nếu chưa có
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = col_name
                cell.font = cell.font.copy(bold=True)
            start_row = 2
        else:
            start_row = last_row + 1
        
        # Ghi dữ liệu
        print(f"\nĐang ghi {len(products)} sản phẩm vào Excel...")
        for row_idx, product in enumerate(products, start=start_row):
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = product.get(col_name, '')
                # Xử lý các giá trị đặc biệt
                if isinstance(value, (list, dict)):
                    value = str(value)
                cell.value = value
        
        # Auto-adjust column widths
        for col_idx in range(1, len(columns) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if row.value:
                        max_length = max(max_length, len(str(row.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(self.excel_file)
        print(f"Đã lưu dữ liệu vào file: {self.excel_file}")
    
    def crawl(self):
        """Main crawl method"""
        try:
            self.setup_driver()
            self.crawl_with_scroll(max_scrolls=20)
            
            if self.all_products:
                self.save_to_excel(self.all_products)
                print(f"\n✅ Hoàn thành! Đã crawl và lưu {len(self.all_products)} sản phẩm vào Excel.")
            else:
                print("\n❌ Không có dữ liệu để lưu.")
        finally:
            if self.driver:
                self.driver.quit()


def main():
    crawler = SamsungCrawlerSelenium()
    crawler.crawl()


if __name__ == "__main__":
    main()

